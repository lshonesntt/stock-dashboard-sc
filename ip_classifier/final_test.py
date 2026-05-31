#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IP 분류 도구 - 최종 완전 테스트 (.xlsx + .xls)
"""

import sys
import os
from datetime import datetime

import openpyxl
import xlrd
import xlwt

SAMPLE_DIR = "/Users/scott/hermes room/ip_classifier"

def read_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row_num in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=row_num, column=col_idx + 1).value for col_idx in range(len(headers))]
        processed = []
        for val in row_vals:
            if isinstance(val, datetime):
                processed.append(val)
            elif isinstance(val, float):
                if val == int(val):
                    processed.append(int(val))
                else:
                    processed.append(val)
            else:
                processed.append(val)
        rows.append(processed)
    return headers, rows


def read_xls(filepath):
    book = xlrd.open_workbook(filepath)
    ws = book.sheet_by_index(0)
    headers = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]
    rows = []
    for row_idx in range(1, ws.nrows):
        row_vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
        processed = []
        for col_idx, val in enumerate(row_vals):
            ctype = ws.cell_type(row_idx, col_idx)
            if ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_tuple(val, book.datemode)
                    processed.append(datetime(*dt))
                except Exception:
                    processed.append(str(val))
            elif isinstance(val, float):
                if val == int(val):
                    processed.append(int(val))
                else:
                    processed.append(val)
            else:
                processed.append(val)
        rows.append(processed)
    return headers, rows


def format_timestamp(val):
    if not val:
        return ""
    if isinstance(val, str):
        return val
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def process_file(filepath, event_type="", enforcer="", ip_param="", admin="", timestamp=""):
    filepath_lower = filepath.lower()
    ext = os.path.splitext(filepath_lower)[1]
    if ext == ".xls":
        headers, rows_data = read_xls(filepath)
    else:
        headers, rows_data = read_xlsx(filepath)

    event_col = None
    for i, h in enumerate(headers):
        if h and "이벤트" in str(h):
            event_col = i
            break

    if event_col is None:
        return None

    new_ip_rows = []
    del_ip_rows = []

    for row_idx, row_vals in enumerate(rows_data):
        if event_col >= len(row_vals):
            continue
        event_val = row_vals[event_col]
        if not event_val or str(event_val).strip() == "":
            continue
        event_str = str(event_val).strip()
        record = {}
        for h_idx, h in enumerate(headers):
            if h_idx < len(row_vals):
                record[str(h)] = row_vals[h_idx]
            else:
                record[str(h)] = ""
        if event_type and str(event_str) != event_type:
            continue
        if enforcer and str(record.get("엔포서명", "") or "") != enforcer:
            continue
        if ip_param and str(record.get("IP", "") or "") != ip_param:
            continue
        if admin and str(record.get("관리자", "") or "") != admin:
            continue
        if timestamp and str(record.get("발생일시", "") or "") != timestamp:
            continue
        if "IP 등록" in event_str or "인증허가" in event_str:
            new_ip_rows.append(record)
        elif "IP 삭제" in event_str or "일정기간 미사용 IP삭제" in event_str:
            del_ip_rows.append(record)

    return {"new_ip": new_ip_rows, "del_ip": del_ip_rows}


def save_output(original_path, new_ip_rows, del_ip_rows, output_dir="."):
    wb = openpyxl.Workbook()
    headers = ["No", "발생일시", "이벤트 종류", "엔포서명", "IP", "관리자"]
    bold_font = openpyxl.styles.Font(bold=True)
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin"),
        right=openpyxl.styles.Side(style="thin"),
        top=openpyxl.styles.Side(style="thin"),
        bottom=openpyxl.styles.Side(style="thin"),
    )

    ws_new = wb.active
    ws_new.title = "신규IP"
    for col, h in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    for idx, row_data in enumerate(new_ip_rows, 1):
        ws_new.cell(row=idx + 1, column=1, value=idx).border = thin_border
        ws_new.cell(row=idx + 1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_new.cell(row=idx + 1, column=2, value=format_timestamp(row_data.get("발생일시", ""))).border = thin_border
        ws_new.cell(row=idx + 1, column=3, value=str(row_data.get("이벤트 종류", "") or "")).border = thin_border
        ws_new.cell(row=idx + 1, column=4, value=row_data.get("엔포서명", "") or "").border = thin_border
        ws_new.cell(row=idx + 1, column=5, value=str(row_data.get("IP", "") or "")).border = thin_border
        ws_new.cell(row=idx + 1, column=6, value=row_data.get("관리자", "") or "").border = thin_border

    ws_del = wb.create_sheet("IP삭제")
    for col, h in enumerate(headers, 1):
        cell = ws_del.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    for idx, row_data in enumerate(del_ip_rows, 1):
        ws_del.cell(row=idx + 1, column=1, value=idx).border = thin_border
        ws_del.cell(row=idx + 1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_del.cell(row=idx + 1, column=2, value=format_timestamp(row_data.get("발생일시", ""))).border = thin_border
        ws_del.cell(row=idx + 1, column=3, value=str(row_data.get("이벤트 종류", "") or "")).border = thin_border
        ws_del.cell(row=idx + 1, column=4, value=row_data.get("엔포서명", "") or "").border = thin_border
        ws_del.cell(row=idx + 1, column=5, value=str(row_data.get("IP", "") or "")).border = thin_border
        ws_del.cell(row=idx + 1, column=6, value=row_data.get("관리자", "") or "").border = thin_border

    max_widths = [10] * len(headers)
    for row_data in new_ip_rows + del_ip_rows:
        vals = [
             "No",
            format_timestamp(row_data.get("발생일시", "")),
            str(row_data.get("이벤트 종류", "") or ""),
            str(row_data.get("엔포서명", "") or ""),
            str(row_data.get("IP", "") or ""),
            str(row_data.get("관리자", "") or ""),
         ]
        for ci, val in enumerate(vals):
            w = len(val.encode("utf-8"))
            if w > max_widths[ci]:
                max_widths[ci] = min(w, 60)

    for ws in [ws_new, ws_del]:
        for ci, max_w in enumerate(max_widths):
            col_letter = ws.cell(row=1, column=ci + 1).column_letter
            ws.column_dimensions[col_letter].width = max_w + 4

    from datetime import datetime, timedelta
    yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
    out_name = f"IP부여_{yesterday_str}.xlsx"
    output_path = os.path.join(output_dir, out_name)
    wb.save(output_path)
    return output_path


# ── 테스트 실행 ──────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 70)
    print("IP 분류 도구 - 최종 완전 테스트 (.xlsx + .xls)")
    print("=" * 70)

    xlsx_files = [
        f"{SAMPLE_DIR}/samples_1.xlsx",
        f"{SAMPLE_DIR}/samples_2.xlsx",
        f"{SAMPLE_DIR}/samples_3.xlsx",
    ]
    xls_files = [
        f"{SAMPLE_DIR}/samples_1.xls",
    ]

    all_passed = True

    for fp in xlsx_files + xls_files:
        if not os.path.exists(fp):
            print("\n" + "x" * 70)
            print(f"  {os.path.basename(fp)} - 파일 없음")
            all_passed = False
            continue

        ext = os.path.splitext(fp)[1]
        print("\n" + "=" * 70)
        print(f"  {os.path.basename(fp)} ({ext})")
        print("=" * 70)

        results = process_file(fp)
        if not results or not results["new_ip"] or not results["del_ip"]:
            print(f"  ERROR: 신규IP: {len(results['new_ip']) if results else 0}, IP삭제: {len(results['del_ip']) if results else 0}")
            all_passed = False
            continue

        saved_path = save_output(fp, results["new_ip"], results["del_ip"], SAMPLE_DIR)

        # 저장된 파일 검증
        verify_wb = openpyxl.load_workbook(saved_path)
        print(f"  시트: {', '.join(verify_wb.sheetnames)}")

        new_ip_count = 0
        del_ip_count = 0

        for sn in verify_wb.sheetnames:
            ws = verify_wb[sn]
            rows = ws.max_row - 1
            if sn == "신규IP":
                new_ip_count = rows
            elif sn == "IP삭제":
                del_ip_count = rows
            print(f"    {sn}: {rows}행")

         # 이벤트 종류 열(3번)의 모든 값 확인
            event_vals = []
            for r in range(2, ws.max_row + 1):
                ev = ws.cell(row=r, column=3).value
                event_vals.append(str(ev) if ev else "")
            print(f"      이벤트 종류: {', '.join(event_vals)}")

         # 검증: 신규IP에는 IP 등록/인증허가만, IP삭제에는 IP삭제/일정기간미사용IP삭제만 있어야 함
        ok = True
        for sn in verify_wb.sheetnames:
            ws = verify_wb[sn]
            for r in range(2, ws.max_row + 1):
                ev = str(ws.cell(row=r, column=3).value or "")
                if sn == "신규IP":
                    if "IP 등록" not in ev and "인증허가" not in ev:
                        print(f"  WARNING: 신규IP 시트에 잘못 분류됨: {ev}")
                        ok = False
                elif sn == "IP삭제":
                    if "IP 삭제" not in ev and "일정기간 미사용 IP삭제" not in ev:
                        print(f"  WARNING: IP삭제 시트에 잘못 분류됨: {ev}")
                        ok = False
            if ok:
                print(f"  검증 통과: {sn} 시트의 이벤트 종류가 모두 올바른 값으로 분류됨")

        print(f"  출력: {saved_path}")

    print("\n" + "=" * 70)
    if all_passed:
        print("모든 테스트 통과!")
    else:
        print("일부 테스트 실패 - 확인 필요")
    print("=" * 70)
