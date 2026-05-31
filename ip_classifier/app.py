#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IP 분류 도구 - 원본 Excel 파일을 읽어 이벤트 종류별로 시트를 분류하는 GUI 앱.
.xls와 .xlsx 파일 모두 지원.
"""

import sys
import os
import re
from datetime import datetime, timedelta

try:
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QLabel, QLineEdit, QPushButton, QFileDialog, QTextEdit,
        QFormLayout, QMessageBox, QGroupBox, QSplitter
    )
    from PyQt5.QtCore import QThread, pyqtSignal, Qt
    from PyQt5.QtGui import QFont
except ImportError:
    print("PyQt5가 필요합니다. pip install PyQt5 후 실행하세요.")
    sys.exit(1)

# Excel 라이브러리
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("openpyxl이 필요합니다. pip install openpyxl 후 실행하세요.")
    sys.exit(1)

try:
    import xlrd
    xlrd_available = True
except ImportError:
    xlrd_available = False


# ── 워커: 엑셀 처리(백그라운드) ──────────────────────────────────────────

class ProcessWorker(QThread):
    progress_signal = pyqtSignal(str, float)
    finished_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)

    def __init__(self, input_path, event_type, enforcer, ip, admin, timestamp, output_dir):
        super().__init__()
        self.input_path = input_path
        self.event_type = event_type
        self.enforcer = enforcer
        self.ip = ip
        self.admin = admin
        self.timestamp = timestamp
        self.output_dir = output_dir

    def run(self):
        try:
            filepath = self.input_path.lower()
            ext = os.path.splitext(filepath)[1]

            # 파일 읽기
            headers = []
            rows_data = []

            if ext == '.xls':
                 # xlrd로 .xls 읽기 (Python 3.9 compatible API)
                if not xlrd_available:
                    raise ImportError("xlrd가 필요합니다. pip install xlrd를 실행하세요.")
                book = xlrd.open_workbook(self.input_path)
                ws = book.sheet_by_index(0)
                if ws.nrows < 1:
                    raise ValueError("파일이 비어있습니다.")
                headers = [str(ws.cell_value(0, col)).strip() for col in range(ws.ncols)]
                for row_idx in range(1, ws.nrows):
                    row_vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
                        # xlrd 날짜/숫자 처리
                    processed = []
                    for val in row_vals:
                        if xlrd.XL_CELL_DATE:
                            try:
                                dt = xlrd.xldate_as_tuple(val, book.datemode)
                                processed.append(datetime(*dt))
                            except Exception:
                                processed.append(str(val))
                        elif isinstance(val, float):
                            if val == int(val):
                                processed.append(int(val))
                            else:
                                processed.append(val)
                        else:
                            processed.append(val)
                    rows_data.append(processed)
            else:
                 # openpyxl로 .xlsx 읽기
                wb = openpyxl.load_workbook(self.input_path)
                ws = wb.active
                if ws.max_row < 1:
                    raise ValueError("파일이 비어있습니다.")
                headers = [cell.value for cell in ws[1]]
                for row_num in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(row=row_num, column=col_idx + 1).value for col_idx in range(len(headers))]
                        # openpyxl 날짜/숫자 처리
                    processed = []
                    for val in row_vals:
                        if isinstance(val, datetime):
                            processed.append(val)
                        elif isinstance(val, float):
                            if val == int(val):
                                processed.append(int(val))
                            else:
                                processed.append(val)
                        else:
                            processed.append(val)
                    rows_data.append(processed)

            # 헤더 확인
            event_col = None
            for i, h in enumerate(headers):
                if h and "이벤트" in str(h):
                    event_col = i
                    break

            if event_col is None:
                raise ValueError("'이벤트 종류' 열을 찾을 수 없습니다.\n원본 파일의 1행(헤더)에 '이벤트 종류' 열이 있는지 확인하세요.")

            new_ip_rows = []
            del_ip_rows = []

            # 데이터 행 검토
            for row_idx, row_vals in enumerate(rows_data):
                if event_col >= len(row_vals):
                    continue
                event_val = row_vals[event_col]
                if not event_val or str(event_val).strip() == "":
                    continue
                event_str = str(event_val).strip()

                # 행 전체를 record로 저장
                record = {}
                for h_idx, h in enumerate(headers):
                    if h_idx < len(row_vals):
                        record[str(h)] = row_vals[h_idx]
                    else:
                        record[str(h)] = ""

                # 조건 A: IP 등록 또는 인증허가
                if "IP 등록" in event_str or "인증허가" in event_str:
                    if self._matches(record, event_str):
                        new_ip_rows.append(record)
                # 조건 B: IP 삭제 또는 일정기간 미사용 IP삭제
                elif "IP 삭제" in event_str or "일정기간 미사용 IP삭제" in event_str:
                    if self._matches(record, event_str):
                        del_ip_rows.append(record)

            # 새 워크북 생성
            out_wb = Workbook()

         # 시트 1: 신규IP (비고 열 포함)
            ws_new = out_wb.active
            ws_new.title = "신규IP"
            self._write_sheet(ws_new, new_ip_rows, has_bigo=True)

             # 시트 2: IP삭제 (비고 열 없음)
            ws_del = out_wb.create_sheet("IP삭제")
            self._write_sheet(ws_del, del_ip_rows, has_bigo=False)

              # 저장
            yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
            out_name = f"IP부여_{yesterday_str}.xlsx"
            out_path = os.path.join(self.output_dir, out_name)
            out_wb.save(out_path)

            self.progress_signal.emit("완료", 1.0)
            self.finished_signal.emit(out_path)

        except Exception as e:
            self.error_signal.emit(str(e))

    def _matches(self, record, event_str):
        """사용자 입력 파라미터와 일치하는지 확인.
        입력 파라미터가 비어있으면 그 열은 필터링 없이 통과.
        """
        if self.event_type and str(event_str) != self.event_type:
            return False
        if self.enforcer and str(record.get("엔포서명", "") or "") != self.enforcer:
            return False
        if self.ip and str(record.get("IP", "") or "") != self.ip:
            return False
        if self.admin and str(record.get("관리자", "") or "") != self.admin:
            return False
        if self.timestamp and str(record.get("발생일시", "") or "") != self.timestamp:
            return False
        return True

    def _write_sheet(self, ws, rows, has_bigo=False):
        """시트에 헤더와 데이터 쓰기.
        has_bigo=True: 신규IP 시트 (비고 열 포함)
        has_bigo=False: IP삭제 시트 (비고 열 없음)
        """
        # 비고 열 유무에 따라 헤더 결정
        if has_bigo:
            headers = ["No", "발생일시", "이벤트 종류", "엔포서명", "IP", "관리자", "비고"]
        else:
            headers = ["No", "발생일시", "이벤트 종류", "엔포서명", "IP", "관리자"]
        
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 데이터 - No/No 열 is already handled at line start but let's include it
        for idx, row_data in enumerate(rows, 1):
            ws.cell(row=idx + 1, column=1, value=idx).border = thin_border
            ws.cell(row=idx + 1, column=1).alignment = Alignment(horizontal='center')

            # 발생일시
            ts = row_data.get("발생일시", "")
            ts_str = self._format_timestamp(ts)
            ws.cell(row=idx + 1, column=2, value=ts_str).border = thin_border

            # 이벤트 종류
            event_val = row_data.get("이벤트 종류", "") or ""
            ws.cell(row=idx + 1, column=3, value=str(event_val)).border = thin_border

            # 엔포서명
            enforcer = row_data.get("엔포서명", "") or ""
            ws.cell(row=idx + 1, column=4, value=enforcer).border = thin_border

            # IP - 숫자일 수 있으니 문자열로
            ip_val = row_data.get("IP", "") or ""
            ws.cell(row=idx + 1, column=5, value=str(ip_val)).border = thin_border

            # 관리자
            admin_val = row_data.get("관리자", "") or ""
            ws.cell(row=idx + 1, column=6, value=admin_val).border = thin_border

            # 비고 (신규IP 시트만)
            if has_bigo:
                bigo_val = row_data.get("비고", "") or ""
                ws.cell(row=idx + 1, column=7, value=str(bigo_val)).border = thin_border

        # 셀 너비 자동 조정
        max_widths = [10] * len(headers)    # 기본 너비
        for row_data in rows:
            vals = [
                str(rows.index(row_data) + 1),    # No
                self._format_timestamp(row_data.get("발생일시", "")),    # 발생일시
                str(row_data.get("이벤트 종류", "") or ""),    # 이벤트 종류
                str(row_data.get("엔포서명", "") or ""),    # 엔포서명
                str(row_data.get("IP", "") or ""),    # IP
                str(row_data.get("관리자", "") or ""),    # 관리자
            ]
            if has_bigo:
                vals.append(str(row_data.get("비고", "") or ""))    # 비고
            for col_idx, val in enumerate(vals):
                w = len(val.encode('utf-8'))    # 한글은 3바이트
                if w > max_widths[col_idx]:
                    max_widths[col_idx] = min(w, 60)    # 최대 60

        for col_idx, max_w in enumerate(max_widths):
            col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
            ws.column_dimensions[col_letter].width = max_w + 4

    def _format_timestamp(self, val):
        """timestamp를 YYYY-MM-DD HH:mm:ss 형식으로 변환."""
        if not val:
            return ""
        if isinstance(val, str):
            return val  # 이미 문자열
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        # openpyxl/xlrd 날짜 객체
        if hasattr(val, 'strftime'):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        return str(val)


# ── 메인 윈도우 ──────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.worker = None
        self.setWindowTitle("IP 분류 도구")
        self.setMinimumSize(750, 600)
        # Fusion 스타일은 macOS PyInstaller 번들에서 오류를 유발하므로 사용 안 함

        # 폰트
        font = QFont("Helvetica Neue", 11)
        font.setWeight(QFont.Normal)
        self.setFont(font)

        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 제목
        title = QLabel("📋 Excel IP 분류 도구")
        title.setFont(QFont("Helvetica Neue", 16, QFont.Bold))
        main_layout.addWidget(title)

        # 파일 선택 그룹
        file_group = QGroupBox("1. 원본 파일")
        file_layout = QHBoxLayout(file_group)
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("원본 Excel 파일을 선택하세요...")
        self.file_path_edit.setReadOnly(True)
        browse_btn = QPushButton("파일 찾기")
        browse_btn.clicked.connect(self._browse_file)
        file_layout.addWidget(self.file_path_edit)
        file_layout.addWidget(browse_btn)
        main_layout.addWidget(file_group)

        # 파라미터 입력 그룹
        param_group = QGroupBox("2. 사용자 입력 파라미터 (선택 — 비워두면 필터링 없이)")
        param_layout = QFormLayout(param_group)

        self.event_type_edit = QLineEdit()
        self.event_type_edit.setPlaceholderText("예: IP 등록")
        self.enforcer_edit = QLineEdit()
        self.enforcer_edit.setPlaceholderText("예: MyEnforcer")
        self.ip_edit = QLineEdit()
        self.ip_edit.setPlaceholderText("예: 192.168.1.100")
        self.admin_edit = QLineEdit()
        self.admin_edit.setPlaceholderText("예: admin")
        self.timestamp_edit = QLineEdit()
        self.timestamp_edit.setPlaceholderText("예: 2026-05-25 14:30:00")

        param_layout.addRow("이벤트 종류:", self.event_type_edit)
        param_layout.addRow("엔포서명:", self.enforcer_edit)
        param_layout.addRow("IP:", self.ip_edit)
        param_layout.addRow("관리자:", self.admin_edit)
        param_layout.addRow("발생일자:", self.timestamp_edit)
        main_layout.addWidget(param_group)

         # 출력 저장 위치 그룹
        save_group = QGroupBox("3. 출력 저장 위치")
        save_layout = QVBoxLayout(save_group)
        self.save_path_edit = QLineEdit()
        self.save_path_edit.setPlaceholderText("출력 파일을 저장할 폴더 경로를 입력하세요...")
        self.save_path_edit.setReadOnly(True)
        browse_save_btn = QPushButton("폴더 찾기")
        browse_save_btn.clicked.connect(self._browse_save_dir)
        save_layout.addWidget(self.save_path_edit)
        save_layout.addWidget(browse_save_btn)
        main_layout.addWidget(save_group)

        # 처리 버튼
        self.process_btn = QPushButton("🚀 처리 시작")
        self.process_btn.setFont(QFont("Helvetica Neue", 12, QFont.Bold))
        self.process_btn.setMinimumHeight(44)
        self.process_btn.clicked.connect(self._process)
        main_layout.addWidget(self.process_btn)

        # 로거
        log_group = QGroupBox("로그 / 결과")
        log_layout = QVBoxLayout(log_group)
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.log_view.setMinimumHeight(150)
        log_layout.addWidget(self.log_view)
        main_layout.addWidget(log_group)

        # 스프링으로 위쪽 밀기
        main_layout.addStretch()

    # ── 이벤트 핸들러 ──────────────────────────────────────────────

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "원본 Excel 파일 선택", "", "Excel Files (*.xlsx *.xls)"
         )
        if path:
            self.file_path_edit.setText(path)
            self._log(f"파일 선택됨: {path}")

    def _browse_save_dir(self):
        path = QFileDialog.getExistingDirectory(
            self, "출력 폴더 선택", ""
         )
        if path:
            self.save_path_edit.setText(path)
            self._log(f"출력 폴더 선택됨: {path}")

    def _log(self, msg):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_view.append(f"[{ts}] {msg}")

    def _process(self):
        input_path = self.file_path_edit.text().strip()
        if not input_path or not os.path.exists(input_path):
            QMessageBox.warning(self, "경고", "올바른 원본 파일 경로를 입력하세요.")
            return

        # 출력 폴더: 지정되지 않으면 원본 파일이 있는 폴더로
        save_dir = self.save_path_edit.text().strip()
        if not save_dir or not os.path.isdir(save_dir):
            save_dir = os.path.dirname(input_path)

        event_type = self.event_type_edit.text().strip()
        enforcer = self.enforcer_edit.text().strip()
        ip = self.ip_edit.text().strip()
        admin = self.admin_edit.text().strip()
        timestamp = self.timestamp_edit.text().strip()

        # 타임스탬프 포맷 검증
        if timestamp:
            try:
                datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                QMessageBox.warning(self, "경고",
                                     "발생일자 형식이 올바르지 않습니다.\nYYYY-MM-DD HH:mm:ss 형식을 사용하세요.")
                return

        self.process_btn.setEnabled(False)
        self._log("처리 시작...")

        # 워커 생성 & 실행
        self.worker = ProcessWorker(
            input_path, event_type, enforcer, ip, admin, timestamp,
            save_dir
          )
        self.worker.progress_signal.connect(self._on_progress)
        self.worker.finished_signal.connect(self._on_finished)
        self.worker.error_signal.connect(self._on_error)
        self.worker.start()

    def _on_progress(self, msg, _pct):
        self._log(msg)

    def _on_finished(self, out_path):
        self.process_btn.setEnabled(True)
        self._log(f"✅ 완료! 출력 파일: {out_path}")
        QMessageBox.information(self, "완료",
                               f"처리 완료!\n\n출력 파일:\n{out_path}")

    def _on_error(self, err):
        self.process_btn.setEnabled(True)
        self._log(f"❌ 오류: {err}")
        QMessageBox.critical(self, "오류", f"처리 중 오류가 발생했습니다:\n{err}")

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            self.worker.terminate()
        event.accept()


# ── 엔트리 ──────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    # Fusion 스타일은 PyInstaller 번들에서 오류를 유발할 수 있음
    try:
        app.setStyle("Fusion")
    except Exception:
        pass
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
