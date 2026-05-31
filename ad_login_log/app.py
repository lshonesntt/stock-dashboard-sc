#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AD 로그인 로그 자동화 도구
- AD 내부망/외부망 시트로로그를 분류하고 정리하는 GUI 앱
- 기존 app.py와 동일한 UI/UX 패턴
- No는 누적 파일 기준 다음 번호부터 시작 (없으면 1부터)
- 회수일자 열 하이라이트 없음 (모든 셀 기본 배경)
"""

import sys
import os
import platform
import re
from datetime import datetime, timedelta
from collections import OrderedDict

try:
    from PyQt5.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QLabel, QLineEdit, QPushButton, QFileDialog, QTextEdit,
        QFormLayout, QMessageBox, QGroupBox, QSplitter, QComboBox
    )
    from PyQt5.QtCore import QThread, pyqtSignal, Qt
    from PyQt5.QtGui import QFont
except ImportError:
    print("PyQt5가 필요합니다. pip install PyQt5 후 실행하세요.")
    sys.exit(1)

# Excel 라이브러리
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("openpyxl이 필요합니다. pip install openpyxl 후 실행하세요.")
    sys.exit(1)

try:
    import xlrd
    xlrd_available = True
except ImportError:
    xlrd_available = False


# ── 날짜 파싱 유틸리티 ──────────────────────────────────────────────


def parse_date_to_datetime(val):
    """다양한 날짜 형식을 YYYY-MM-DD HH:mm:ss 파싱된 datetime 객체로 변환."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        # xlrd 날짜 숫자 처리
        try:
            dt = xlrd.xldate_as_tuple(val, 0)
            return datetime(*dt)
        except Exception:
            return None
    val_str = str(val).strip()
    if not val_str:
        return None
    # dateutil_parser 시도 (유연한 파싱)
    try:
        from dateutil import parser as dtp
        dt = dtp.parse(val_str, dayfirst=False)
        return dt
    except Exception:
        pass
    # 패턴 1: YYYY/MM/DD HH:MM:SS 또는 YYYY-MM-DD HH:MM:SS
    m = re.match(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})\s+(\d{1,2}):(\d{2}):?(\d{2})?', val_str)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                        int(m.group(4)), int(m.group(5)), int(m.group(6) if m.group(6) else 0))
    # 패턴 2: MM-DD-YYYY HH:MM:SS
    m = re.match(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})\s+(\d{1,2}):(\d{2}):?(\d{2})?', val_str)
    if m:
        return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2)),
                        int(m.group(4)), int(m.group(5)), int(m.group(6) if m.group(6) else 0))
    # 패턴 3: YYYY/MM/DD (시간 없음)
    m = re.match(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$', val_str)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # 패턴 4: MM-DD-YYYY (시간 없음)
    m = re.match(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', val_str)
    if m:
        return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    return val_str


def format_datetime(val):
    """datetime 객체를 YYYY-MM-DD HH:mm:ss 형식으로 변환."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, str):
        dt = parse_date_to_datetime(val)
        if isinstance(dt, datetime):
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        return val
    if hasattr(val, 'strftime'):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def get_max_no(rows_data):
    """기존 No 열에서 최대 정수값을 찾음."""
    max_no = 0
    for record in rows_data:
        no_val = record.get("no", record.get("No", ""))
        if no_val is not None:
            try:
                num = int(float(str(no_val).strip()))
                if num > max_no:
                    max_no = num
            except (ValueError, TypeError):
                pass
    return max_no


def get_windows_font():
    """윈도우에서 한글이 잘 보이는 폰트 반환"""
    try:
        return QFont("Malgun Gothic", 11)
    except Exception:
        return QFont("굴림", 11)


def get_mac_font():
    return QFont("Helvetica Neue", 11)


def get_base_font():
    system = platform.system()
    if "Darwin" in system:
        return get_mac_font()
    return get_windows_font()


def get_title_font(size=16, bold=True):
    system = platform.system()
    if "Darwin" in system:
        return QFont("Helvetica Neue", size, QFont.Bold if bold else QFont.Normal)
    return QFont("Malgun Gothic", size, QFont.Bold if bold else QFont.Normal)


# ── 워커: 엑셀 처리(백그라운드) ───────────────────────────────────────


class ProcessWorker(QThread):
    progress_signal = pyqtSignal(str, float)
    finished_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)

    def __init__(self, input_path, cumulative_path, output_dir, date_col_index):
        super().__init__()
        self.input_path = input_path
        self.cumulative_path = cumulative_path
        self.output_dir = output_dir
        self.date_col_index = date_col_index    # 회수일자 열 인덱스 (0-based)

    def _read_cumulative_max_no(self, cumul_path):
        """누적 파일에서 최대 No 값을 읽음."""
        if not cumul_path or not os.path.exists(cumul_path):
            return 0
        max_no = 0
        try:
            ext = os.path.splitext(cumul_path.lower())[1]
            if ext == '.xls':
                if not xlrd_available:
                    return 0
                book = xlrd.open_workbook(cumul_path)
                ws = book.sheet_by_index(0)
                for row_idx in range(1, ws.nrows):
                    val = ws.cell_value(row_idx, 0)
                    if val is not None:
                        try:
                            num = int(float(str(val).strip()))
                            if num > max_no:
                                max_no = num
                        except (ValueError, TypeError):
                            pass
                return max_no
            elif ext in ('.xlsx', '.xlsm'):
                wb = openpyxl.load_workbook(cumul_path, read_only=True)
                ws = wb.active
                for row_num in range(2, ws.max_row + 1):
                    val = ws.cell(row=row_num, column=1).value
                    if val is not None:
                        try:
                            num = int(float(str(val).strip()))
                            if num > max_no:
                                max_no = num
                        except (ValueError, TypeError):
                            pass
                wb.close()
                return max_no
            else:
                return 0
        except Exception:
            return 0

    def _read_cumulative_keys(self, cumul_path, sheet_name):
        """누적 파일에서 (회수일자, 회수내용) 키 집합을 읽음."""
        if not cumul_path or not os.path.exists(cumul_path):
            return set()
        keys = set()
        try:
            ext = os.path.splitext(cumul_path.lower())[1]
            if ext == '.xls':
                if not xlrd_available:
                    return keys
                book = xlrd.open_workbook(cumul_path)
                target_ws = None
                for i in range(book.nsheets):
                    ws = book.sheet_by_index(i)
                    if ws.title == sheet_name:
                        target_ws = ws
                        break
                if target_ws is None:
                    target_ws = book.sheet_by_index(0)
                for row_idx in range(1, target_ws.nrows):
                    date_val = str(target_ws.cell_value(row_idx, 1)).strip()
                    content_val = str(target_ws.cell_value(row_idx, 2)).strip()
                    if date_val and content_val:
                        keys.add((date_val, content_val))
            elif ext in ('.xlsx', '.xlsm'):
                wb = openpyxl.load_workbook(cumul_path, read_only=True)
                target_ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
                for row_num in range(2, target_ws.max_row + 1):
                    date_val = str(target_ws.cell(row=row_num, column=2).value or "").strip()
                    content_val = str(target_ws.cell(row=row_num, column=3).value or "").strip()
                    if date_val and content_val:
                        keys.add((date_val, content_val))
                wb.close()
        except Exception:
            pass
        return keys

    def run(self):
        try:
            filepath = self.input_path.lower()
            ext = os.path.splitext(filepath)[1]

            # 파일 읽기
            headers = []
            rows_data = []

            if ext == '.xls':
                if not xlrd_available:
                    raise ImportError("xlrd가 필요합니다. pip install xlrd를 실행하세요.")
                book = xlrd.open_workbook(self.input_path)
                ws = book.sheet_by_index(0)
                if ws.nrows < 1:
                    raise ValueError("파일이 비어있습니다.")
                headers = [str(ws.cell_value(0, col)).strip() for col in range(ws.ncols)]
                for row_idx in range(1, ws.nrows):
                    row_vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
                    processed = []
                    for val in row_vals:
                        if xlrd.XL_CELL_DATE:
                            try:
                                dt = xlrd.xldate_as_tuple(val, book.datemode)
                                processed.append(datetime(*dt))
                            except Exception:
                                processed.append(str(val))
                        elif isinstance(val, float):
                            if val == int(val):
                                processed.append(int(val))
                            else:
                                processed.append(val)
                        else:
                            processed.append(val)
                    rows_data.append(processed)
            elif ext in ('.xlsx', '.xlsm'):
                wb = openpyxl.load_workbook(self.input_path)
                ws = wb.active
                if ws.max_row < 1:
                    raise ValueError("파일이 비어있습니다.")
                headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
                for row_num in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(row=row_num, column=col_idx + 1).value for col_idx in range(len(headers))]
                    processed = []
                    for val in row_vals:
                        if isinstance(val, datetime):
                            processed.append(val)
                        elif isinstance(val, float):
                            if val == int(val):
                                processed.append(int(val))
                            else:
                                processed.append(val)
                        else:
                            processed.append(val)
                    rows_data.append(processed)
            elif ext == '.csv':
                import csv
                with open(self.input_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    raw_headers = next(reader)
                    headers = [h.strip() for h in raw_headers]
                    for row_vals in reader:
                        processed = []
                        for val in row_vals:
                            processed.append(val)
                        rows_data.append(processed)
            else:
                raise ValueError(f"지원하지 않는 파일 형식입니다: {ext}")

            # 헤더 인덱스 찾기
            no_col = None
            date_col = None
            content_col = None
            reason_col = None

            for i, h in enumerate(headers):
                h_lower = str(h).strip().lower() if h else ""
                if h_lower in ('no', '번호', 'no.'):
                    no_col = i
                if '회수일자' in str(h) or '로그일자' in str(h) or '날짜' in str(h):
                    date_col = i
                if '회수내용' in str(h) or '내용' in str(h):
                    content_col = i
                if '회수사유' in str(h) or '분류' in str(h) or '사유' in str(h):
                    reason_col = i

            # 사용자 지정 date_col_index 우선 사용
            if self.date_col_index is not None and self.date_col_index >= 0 and self.date_col_index < len(headers):
                date_col = self.date_col_index
                self.progress_signal.emit(
                    f"사용자가 선택한 회수일자 열: {headers[self.date_col_index]}", 0.15
                )
            elif date_col is None:
                raise ValueError("'회수일자'(또는 '로그일자') 열을 찾을 수 없습니다.")

            if content_col is None:
                raise ValueError("'회수내용'열을 찾을 수 없습니다.")
            if reason_col is None:
                raise ValueError("'회수사유'(또는 '분류') 열을 찾을 수 없습니다.")

            # 행 데이터를 record 딕셔너리로 변환
            all_records = []
            for row_idx, row_vals in enumerate(rows_data):
                if no_col is not None and no_col >= len(row_vals):
                    continue
                if date_col >= len(row_vals):
                    continue
                if content_col >= len(row_vals):
                    continue
                if reason_col >= len(row_vals):
                    continue

                # 빈 행 건너뛰기
                if not any(row_vals[i] for i in [date_col, content_col, reason_col]):
                    continue

                record = {}
                if no_col is not None and no_col < len(row_vals):
                    record["no"] = row_vals[no_col]
                record["회수일자"] = row_vals[date_col]
                record["회수내용"] = row_vals[content_col]
                record["회수사유"] = str(row_vals[reason_col]).strip() if row_vals[reason_col] else ""
                record["_original_index"] = row_idx + 2
                all_records.append(record)

            # ── 회수사유별 필터링 및 중복 제거 ───────────────────────────
            allowed_internal = {"사원퇴사", "일정기간 로그인없음", "패스워드 변경안함"}

            internal_records = []
            external_records = []
            skipped_reasons = set()
            seen_internal = set()
            seen_external = set()
            skip_count = 0
            dup_count = 0

            for record in all_records:
                reason = record["회수사유"]
                # 허용되지 않은 사유는 스킵
                if reason not in allowed_internal:
                    skipped_reasons.add(reason)
                    skip_count += 1
                    continue

                # 회수사유 + 회수내용으로 중복 체크
                dedup_key = (str(record["회수일자"]), str(record["회수내용"]))

                if reason == "사원퇴사":
                     # 사원퇴사는 양쪽 시트에 모두 - 서로 다른 객체로 복사
                    if dedup_key in seen_external:
                        dup_count += 1
                        continue
                    seen_external.add(dedup_key)
                    external_records.append(dict(record))
                     # 내부망에도 포함 (원본 그대로)
                    if dedup_key not in seen_internal:
                        seen_internal.add(dedup_key)
                        internal_records.append(record)
                elif reason in ("일정기간 로그인없음", "패스워드 변경안함"):
                    if dedup_key in seen_internal:
                        dup_count += 1
                        continue
                    seen_internal.add(dedup_key)
                    internal_records.append(record)

            # ── 누적 파일에서 기존 No 최대값 파악 (선택) ────────────────
            existing_max_no = self._read_cumulative_max_no(self.cumulative_path)
            if existing_max_no > 0:
                self.progress_signal.emit(f"누적 파일에서 기존 No 최대값: {existing_max_no} 가져옴", 0.05)
            else:
                self.progress_signal.emit("누적 파일 없음 — No는 1부터 시작합니다", 0.05)

            # ── No 값 정수 할당 (누적 — 마지막 정수 다음부터) ─────────────
            internal_no_counter = existing_max_no
            external_no_counter = existing_max_no

            for rec in internal_records:
                internal_no_counter += 1
                rec["no"] = internal_no_counter
            for rec in external_records:
                external_no_counter += 1
                rec["no"] = external_no_counter

            self.progress_signal.emit(f"AD내부망: {internal_no_counter}개 행", 0.5)
            self.progress_signal.emit(f"AD외부망: {external_no_counter}개 행", 0.6)

            # ── 새 워크북 생성 ────────────────────────────────────────────
            out_wb = Workbook()

            # 시트 1: AD내부망
            ws_internal = out_wb.active
            ws_internal.title = "AD내부망"
            self._write_sheet(ws_internal, internal_records)

            # 시트 2: AD외부망
            ws_external = out_wb.create_sheet("AD외부망")
            self._write_sheet(ws_external, external_records)

            # ── 저장 ─────────────────────────────────────────────────────
            yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
            out_name = f"AD로그인 로그_{yesterday_str}.xlsx"
            out_path = os.path.join(self.output_dir, out_name)
            out_wb.save(out_path)

            # 결과 정보
            result_info = (
                f"처리 완료!\n\n"
                f"AD내부망: {len(internal_records)}개 행 "
                f"(Skip: {skip_count}개, 중복: {dup_count}개)\n"
                f"AD외부망: {len(external_records)}개 행 "
                f"(Skip: {skip_count}개, 중복: {dup_count}개)\n"
                f"출력 파일: {out_path}\n"
            )
            if skipped_reasons:
                result_info += f"\n스킵된 사유: {', '.join(sorted(skipped_reasons))}"

            self.progress_signal.emit("완료", 1.0)
            self.finished_signal.emit(out_path)
            self.finished_signal.emit(result_info)

        except Exception as e:
            self.error_signal.emit(str(e))

    def _write_sheet(self, ws, rows):
        """시트에 헤더와 데이터 쓰기. (하이라이트 없음 - 모든 셀 기본 배경)"""
        headers = ["No", "회수일자", "회수내용", "회수사유"]
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 데이터
        for idx, row_data in enumerate(rows, 1):
            # No
            no_val = row_data.get("no", "")
            try:
                ws.cell(row=idx + 1, column=1,
                        value=int(no_val)).border = thin_border
            except (ValueError, TypeError):
                ws.cell(row=idx + 1, column=1,
                        value=str(no_val)).border = thin_border
            ws.cell(row=idx + 1, column=1).alignment = Alignment(horizontal='center')

            # 회수일자 (노란색 배경 없음)
            date_val = row_data.get("회수일자", "")
            date_str = format_datetime(date_val)
            cell_date = ws.cell(row=idx + 1, column=2)
            cell_date.value = date_str
            cell_date.border = thin_border

            # 회수내용
            content_val = row_data.get("회수내용", "") or ""
            cell_content = ws.cell(row=idx + 1, column=3)
            cell_content.value = str(content_val)
            cell_content.border = thin_border

            # 회수사유
            reason_val = row_data.get("회수사유", "") or ""
            cell_reason = ws.cell(row=idx + 1, column=4)
            cell_reason.value = str(reason_val)
            cell_reason.border = thin_border

        # 셀 너비 자동 조정
        max_widths = [8] * len(headers)
        for row_data in rows:
            vals = [
                str(row_data.get("no", "")),
                format_datetime(row_data.get("회수일자", "")),
                str(row_data.get("회수내용", "") or ""),
                str(row_data.get("회수사유", "") or ""),
            ]
            for col_idx, val in enumerate(vals):
                w = min(len(val) * 1.5, 50)
                if w > max_widths[col_idx]:
                    max_widths[col_idx] = int(w)

        for col_idx, max_w in enumerate(max_widths):
            col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
            ws.column_dimensions[col_letter].width = max_w + 4

        # 회수일자 열 너비 늘리기 (YYYY-MM-DD HH:mm:ss 형식)
        ws.column_dimensions['B'].width = 25


# ── 메인 윈도우 ──────────────────────────────────────────────────────


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.worker = None
        self.setWindowTitle("AD 로그인 로그 자동화")
        self.setMinimumSize(800, 620)
        self.headers = []
        self.date_col_index = None

        # 폰트
        font = get_base_font()
        font.setWeight(QFont.Normal)
        self.setFont(font)

        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 제목
        title = QLabel("📋 AD 로그인 로그 자동화 도구")
        title.setFont(get_title_font(16, True))
        main_layout.addWidget(title)

        # 파일 선택 그룹 (원본)
        file_group = QGroupBox("1. 원본 파일 (처리할 AD 로그)")
        file_layout = QHBoxLayout(file_group)
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("원본 Excel/CSV 파일을 선택하세요...")
        self.file_path_edit.setReadOnly(True)
        self.file_path_edit.setMinimumHeight(32)
        browse_btn = QPushButton("파일 찾기")
        browse_btn.clicked.connect(self._browse_file)
        file_layout.addWidget(self.file_path_edit)
        file_layout.addWidget(browse_btn)
        main_layout.addWidget(file_group)

        # 누적 파일 선택 그룹 (선택)
        cumul_group = QGroupBox("2. 누적 파일 (선택)")
        cumul_layout = QHBoxLayout(cumul_group)
        self.cumulative_path_edit = QLineEdit()
        self.cumulative_path_edit.setPlaceholderText(
            "누적 파일 선택 (No 이어받기/색칠 기준 - 선택사항)"
        )
        self.cumulative_path_edit.setReadOnly(True)
        self.cumulative_path_edit.setMinimumHeight(32)
        browse_cumul_btn = QPushButton("파일 찾기")
        browse_cumul_btn.clicked.connect(self._browse_cumul_file)
        cumul_layout.addWidget(self.cumulative_path_edit)
        cumul_layout.addWidget(browse_cumul_btn)
        main_layout.addWidget(cumul_group)

        # 출력 저장 위치 그룹
        save_group = QGroupBox("3. 출력 저장 위치")
        save_layout = QVBoxLayout(save_group)
        self.save_path_edit = QLineEdit()
        self.save_path_edit.setPlaceholderText(
            "출력 파일을 저장할 폴더 경로를 입력하세요..."
        )
        self.save_path_edit.setReadOnly(True)
        self.save_path_edit.setMinimumHeight(32)
        browse_save_btn = QPushButton("폴더 찾기")
        browse_save_btn.clicked.connect(self._browse_save_dir)
        save_layout.addWidget(self.save_path_edit)
        save_layout.addWidget(browse_save_btn)
        main_layout.addWidget(save_group)

        # 회수일자 열 선택
        date_col_group = QGroupBox(
            "4. 회수일자 열 선택 (선택 사항 - 자동 인식됨)"
        )
        date_col_layout = QHBoxLayout(date_col_group)
        date_col_layout.addWidget(QLabel(
            "원본 파일에서 회수일자가 포함된 열:"
        ))
        self.date_col_combo = QComboBox()
        self.date_col_combo.setMinimumHeight(32)
        self.date_col_combo.addItems(["[자동 감지]"])
        self.date_col_combo.currentIndexChanged.connect(
            self._on_date_col_changed
        )
        date_col_layout.addWidget(self.date_col_combo)
        date_col_layout.addStretch()
        main_layout.addWidget(date_col_group)

        # 처리 버튼
        self.process_btn = QPushButton("🚀 처리 시작")
        self.process_btn.setFont(get_title_font(12, True))
        self.process_btn.setMinimumHeight(44)
        self.process_btn.clicked.connect(self._process)
        main_layout.addWidget(self.process_btn)

        # 로거
        log_group = QGroupBox("로그 / 결과")
        log_layout = QVBoxLayout(log_group)
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.log_view.setMinimumHeight(150)
        log_layout.addWidget(self.log_view)
        main_layout.addWidget(log_group)

        # 스프링으로 위쪽 밀기
        main_layout.addStretch()

    # ── 이벤트 핸들러 ──────────────────────────────────────────────

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "원본 Excel/CSV 파일 선택", "",
            "All Supported Files (*.xlsx *.xlsm *.xls *.csv);;"
            "Excel Files (*.xlsx *.xlsm *.xls);;"
            "CSV Files (*.csv);;All Files (*)"
        )
        if path:
            path = path.replace("\\", "/")
            self.file_path_edit.setText(path)
            self._log(f"파일 선택됨: {path}")
            self._load_headers(path)
            self._check_ready()

    def _load_headers(self, file_path):
        """원본 파일의 헤더를 읽어서 date_col_combo에 추가"""
        self.date_col_combo.clear()
        self.date_col_combo.addItem("[자동 감지]")
        self.date_col_index = None

        ext = os.path.splitext(file_path.lower())[1]
        try:
            if ext == '.xls':
                if not xlrd_available:
                    return
                book = xlrd.open_workbook(file_path)
                ws = book.sheet_by_index(0)
                headers = [
                    str(ws.cell_value(0, col)).strip()
                    for col in range(ws.ncols)
                ]
            elif ext in ('.xlsx', '.xlsm'):
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                headers = [
                    cell.value if cell.value is not None else ""
                    for cell in ws[1]
                ]
                wb.close()
            elif ext == '.csv':
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    import csv
                    reader = csv.reader(f)
                    headers = [h.strip() for h in next(reader)]
            else:
                return

            self.headers = headers
            for i, h in enumerate(headers):
                self.date_col_combo.addItem(f"{i+1}: {h}")

            # 자동으로 회수일자 열 찾기
            for i, h in enumerate(headers):
                if ('회수일자' in str(h) or '로그일자' in str(h)
                        or '날짜' in str(h)):
                    self.date_col_index = i
                    self.date_col_combo.setCurrentIndex(
                        i + 1    # +1 because "[자동 감지]" is index 0
                    )
                    self._log(f"자동 감지: '{h}'이(가) 회수일자 열")
                    break

        except Exception as e:
            self._log(f"헤더 로드 오류: {e}")

    def _browse_cumul_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "누적 Excel 파일 선택", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if path:
            path = path.replace("\\", "/")
            self.cumulative_path_edit.setText(path)
            self._log(f"누적 파일 선택됨: {path}")
            self._check_ready()

    def _on_date_col_changed(self, index):
        """회수일자 열 변경 시"""
        if index == 0:
            self.date_col_index = None    # 자동 감지
        else:
            self.date_col_index = index - 1    # 0-based 인덱스

    def _check_ready(self):
        """원본 파일이 선택되면 처리 버튼 활성화 (누적 파일 없음)"""
        input_path = self.file_path_edit.text().strip()
        if input_path and os.path.exists(input_path):
            self.process_btn.setEnabled(True)
        else:
            self.process_btn.setEnabled(False)

    def _browse_save_dir(self):
        path = QFileDialog.getExistingDirectory(
            self, "출력 폴더 선택", ""
        )
        if path:
            path = path.replace("\\", "/")
            self.save_path_edit.setText(path)
            self._log(f"출력 폴더 선택됨: {path}")

    def _log(self, msg):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_view.append(f"[{ts}] {msg}")

    def _process(self):
        input_path = self.file_path_edit.text().strip()
        if not input_path or not os.path.exists(input_path):
            path = input_path.replace("\\", "/")
            QMessageBox.warning(
                self, "경고",
                f"올바른 원본 파일 경로를 입력하세요.\n{path}"
            )
            return

        # 누적 파일 (선택)
        cumulative_path = self.cumulative_path_edit.text().strip()
        if cumulative_path and not os.path.exists(cumulative_path):
            cumulative_path = cumulative_path.replace("\\", "/")
            QMessageBox.warning(
                self, "경고",
                f"누적 파일이 존재하지 않습니다:\n{cumulative_path}"
            )
            return

        # 출력 폴더: 지정되지 않으면 원본 파일이 있는 폴더로
        save_dir = self.save_path_edit.text().strip()
        if not save_dir or not os.path.isdir(save_dir):
            save_dir = os.path.dirname(input_path)

        # 파일 존재 확인
        if not os.path.isfile(input_path):
            QMessageBox.warning(self, "경고", "파일이 존재하지 않습니다.")
            return

        self.process_btn.setEnabled(False)
        self._log("처리 시작...")

        # 워커 생성 & 실행
        self.worker = ProcessWorker(
            input_path, cumulative_path, save_dir, self.date_col_index
        )
        self.worker.progress_signal.connect(self._on_progress)
        self.worker.finished_signal.connect(self._on_finished)
        self.worker.error_signal.connect(self._on_error)
        self.worker.start()

    def _on_progress(self, msg, _pct):
        self._log(msg)

    def _on_finished(self, result):
        self.process_btn.setEnabled(True)
        if "처리 완료" in result:
            self._log(result)
            QMessageBox.information(self, "완료", result)
        else:
            self._log(f"✅ 완료! 출력 파일: {result}")
            QMessageBox.information(
                self, "완료",
                f"처리 완료!\n\n출력 파일:\n{result}"
            )

    def _on_error(self, err):
        self.process_btn.setEnabled(True)
        self._log(f"❌ 오류: {err}")
        QMessageBox.critical(
            self, "오류",
            f"처리 중 오류가 발생했습니다:\n{err}"
        )

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            self.worker.terminate()
        event.accept()


# ── 엔트리 ──────────────────────────────────────────────────────────


def main():
    app = QApplication(sys.argv)
    try:
        app.setStyle("Fusion")
    except Exception:
        pass
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
