#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
샘플 엑셀 파일 3개 생성 — IP 분류 도구 테스트용
"""

import random
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

SAMPLES = [
    {
        "filename": "/Users/scott/hermes room/ip_classifier/samples_1.xlsx",
        "desc": "기본 패턴: IP 등록, IP 삭제 혼합",
        "rows": [
            {"이벤트 종류": "IP 등록", "엔포서명": "firewall01", "IP": "192.168.1.10", "관리자": "admin01", "발생일시": "2026-05-20 09:15:30"},
            {"이벤트 종류": "IP 삭제", "엔포서명": "firewall01", "IP": "10.0.0.55", "관리자": "admin02", "발생일시": "2026-05-20 10:22:45"},
            {"이벤트 종류": "로그인", "엔포서명": "firewall01", "IP": "172.16.0.3", "관리자": "admin01", "발생일시": "2026-05-20 11:05:12"},
            {"이벤트 종류": "인증허가", "엔포서명": "firewall02", "IP": "192.168.2.50", "관리자": "admin03", "발생일시": "2026-05-21 08:30:00"},
            {"이벤트 종류": "IP 삭제", "엔포서명": "firewall01", "IP": "10.0.0.100", "관리자": "admin02", "발생일시": "2026-05-21 14:17:33"},
            {"이벤트 종류": "파일 업로드", "엔포서명": "web01", "IP": "192.168.5.80", "관리자": "admin01", "발생일시": "2026-05-21 16:45:22"},
            {"이벤트 종류": "IP 등록", "엔포서명": "firewall02", "IP": "172.16.1.200", "관리자": "admin01", "발생일시": "2026-05-22 07:00:10"},
            {"이벤트 종류": "IP 등록", "엔포서명": "firewall01", "IP": "10.10.0.15", "관리자": "admin04", "발생일시": "2026-05-22 09:33:41"},
            {"이벤트 종류": "시스템 경고", "엔포서명": "monitor01", "IP": "192.168.0.1", "관리자": "admin01", "발생일시": "2026-05-22 12:00:00"},
            {"이벤트 종류": "일정기간 미사용 IP삭제", "엔포서명": "firewall01", "IP": "10.0.0.200", "관리자": "admin02", "발생일시": "2026-05-23 02:00:15"},
        ],
    },
    {
        "filename": "/Users/scott/hermes room/ip_classifier/samples_2.xlsx",
        "desc": "인증허가 중심 + IP삭제 혼합",
        "rows": [
            {"이벤트 종류": "인증허가", "엔포서명": "auth01", "IP": "192.168.10.5", "관리자": "sec_admin", "발생일시": "2026-05-18 06:00:00"},
            {"이벤트 종류": "인증허가", "엔포서명": "auth01", "IP": "192.168.10.12", "관리자": "sec_admin", "발생일시": "2026-05-18 06:15:22"},
            {"이벤트 종류": "IP 등록", "엔포서명": "auth02", "IP": "10.20.30.40", "관리자": "net_admin", "발생일시": "2026-05-18 07:30:00"},
            {"이벤트 종류": "접속 거부", "엔포서명": "auth01", "IP": "203.0.113.50", "관리자": "sec_admin", "발생일시": "2026-05-18 08:45:11"},
            {"이벤트 종류": "IP 삭제", "엔포서명": "auth02", "IP": "192.168.10.9", "관리자": "net_admin", "발생일시": "2026-05-18 09:00:00"},
            {"이벤트 종류": "인증허가", "엔포서명": "auth01", "IP": "192.168.10.77", "관리자": "sec_admin", "발생일시": "2026-05-19 10:20:33"},
            {"이벤트 종류": "파일 다운로드", "엔포서명": "web02", "IP": "172.20.0.5", "관리자": "web_admin", "발생일시": "2026-05-19 11:05:44"},
            {"이벤트 종류": "IP 등록", "엔포서명": "auth01", "IP": "10.20.30.100", "관리자": "sec_admin", "발생일시": "2026-05-19 14:00:00"},
            {"이벤트 종류": "일정기간 미사용 IP삭제", "엔포서명": "auth02", "IP": "192.168.10.20", "관리자": "net_admin", "발생일시": "2026-05-20 02:30:00"},
            {"이벤트 종류": "인증허가", "엔포서명": "auth01", "IP": "192.168.10.88", "관리자": "sec_admin", "발생일시": "2026-05-20 08:15:00"},
            {"이벤트 종류": "IP 삭제", "엔포서명": "auth01", "IP": "10.20.30.1", "관리자": "sec_admin", "발생일시": "2026-05-20 10:00:00"},
            {"이벤트 종류": "IP 등록", "엔포서명": "auth02", "IP": "172.20.0.55", "관리자": "net_admin", "발생일시": "2026-05-20 13:45:22"},
        ],
    },
    {
        "filename": "/Users/scott/hermes room/ip_classifier/samples_3.xlsx",
        "desc": "다양한 이벤트 혼합 + 빈 셀 포함",
        "rows": [
            {"이벤트 종류": "IP 등록", "엔포서명": "fw01", "IP": "10.0.1.100", "관리자": "admin_a", "발생일시": "2026-05-15 10:00:00"},
            {"이벤트 종류": "로그인", "엔포서명": "fw01", "IP": "192.168.50.1", "관리자": "admin_b", "발생일시": "2026-05-15 10:30:00"},
            {"이벤트 종류": "IP 삭제", "엔포서명": "fw02", "IP": "10.0.2.200", "관리자": "admin_a", "발생일시": "2026-05-15 11:00:00"},
            {"이벤트 종류": "인증허가", "엔포서명": "fw01", "IP": "172.30.0.15", "관리자": "admin_c", "발생일시": "2026-05-15 11:30:00"},
            {"이벤트 종류": "시스템 업데이트", "엔포서명": "fw01", "IP": "10.0.0.1", "관리자": "admin_a", "발생일시": "2026-05-15 12:00:00"},
            {"이벤트 종류": "", "엔포서명": "fw02", "IP": "192.168.99.1", "관리자": "admin_b", "발생일시": "2026-05-15 12:30:00"},
            {"이벤트 종류": "IP 등록", "엔포서명": "fw02", "IP": "10.0.3.50", "관리자": "admin_c", "발생일시": "2026-05-16 08:00:00"},
            {"이벤트 종류": "일정기간 미사용 IP삭제", "엔포서명": "fw01", "IP": "172.30.0.99", "관리자": "admin_a", "발생일시": "2026-05-16 02:00:00"},
            {"이벤트 종류": "인증허가", "엔포서명": "fw02", "IP": "192.168.50.200", "관리자": "admin_b", "발생일시": "2026-05-16 09:15:00"},
            {"이벤트 종류": "경고", "엔포서명": "fw01", "IP": "10.0.0.5", "관리자": "admin_a", "발생일시": "2026-05-16 10:00:00"},
            {"이벤트 종류": "IP 삭제", "엔포서명": "fw02", "IP": "172.30.0.50", "관리자": "admin_b", "발생일시": "2026-05-16 14:30:00"},
            {"이벤트 종류": "IP 등록", "엔포서명": "fw01", "IP": "10.0.4.75", "관리자": "admin_c", "발생일시": "2026-05-17 07:45:00"},
            {"이벤트 종류": "인증허가", "엔포서명": "", "IP": "192.168.60.10", "관리자": "admin_a", "발생일시": "2026-05-17 08:00:00"},
            {"이벤트 종류": "시스템 경고", "엔포서명": "fw02", "IP": "10.0.0.2", "관리자": "admin_b", "발생일시": "2026-05-17 09:30:00"},
            {"이벤트 종류": "IP 등록", "엔포서명": "fw01", "IP": "10.0.5.30", "관리자": "admin_a", "발생일시": "2026-05-17 11:00:00"},
        ],
    },
]

def write_header(ws):
    headers = ["No", "발생일시", "엔포서명", "IP", "관리자", "이벤트 종류"]
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

def write_data(ws, rows, start_row=2):
    for idx, row_data in enumerate(rows):
        r = start_row + idx
        ws.cell(row=r, column=1, value=idx + 1)   # No
        for col_idx, header in enumerate(["발생일시", "엔포서명", "IP", "관리자", "이벤트 종류"], 2):
            val = row_data.get(header, "")
            if val is None:
                val = ""
            ws.cell(row=r, column=col_idx, value=val)

def make_sample(filename, desc, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "원본"
    write_header(ws)
    write_data(ws, rows)

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(len(col_letter), max_len + 4)

    wb.save(filename)
    print(f"  ✓ {filename}  ({desc})")
    print(f"    원본 행 수: {len(rows)}")

    # 분류 결과 예측
    new_ip = [r for r in rows if "IP 등록" in str(r["이벤트 종류"]) or "인증허가" in str(r["이벤트 종류"])]
    del_ip = [r for r in rows if "IP 삭제" in str(r["이벤트 종류"]) or "일정기간 미사용 IP삭제" in str(r["이벤트 종류"])]
    others = [r for r in rows if not (
        "IP 등록" in str(r["이벤트 종류"]) or "인증허가" in str(r["이벤트 종류"]) or
        "IP 삭제" in str(r["이벤트 종류"]) or "일정기간 미사용 IP삭제" in str(r["이벤트 종류"])
    )]
    print(f"    신규IP: {len(new_ip)}, IP삭제: {len(del_ip)}, 기타: {len(others)}")

if __name__ == "__main__":
    import os
    os.makedirs("/Users/scott/hermes room/ip_classifier/samples", exist_ok=True)
    print("샘플 엑셀 파일 3개 생성 중...")
    for s in SAMPLES:
        write_sample = lambda: make_sample(s["filename"], s["desc"], s["rows"])
        try:
            write_sample()
        except Exception as e:
            print(f"  ✗ {s['filename']}: {e}")
    print("완료!")
