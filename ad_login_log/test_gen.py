#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AD 로그인 로그용 샘플 데이터 생성"""

import openpyxl
from openpyxl.styles import Font

# 누적 파일 (기존 데이터, No 1~5)
wb_cumul = openpyxl.Workbook()
ws_cumul = wb_cumul.active
ws_cumul.title = "AD내부망"
headers = ["No", "회수일자", "회수내용", "회수사유"]
bold = Font(bold=True)
for col, h in enumerate(headers, 1):
    cell = ws_cumul.cell(row=1, column=col, value=h)
    cell.font = bold

# 데이터 (No 1~5)
for idx, data in enumerate([
    (1, "2026-05-20 09:15:30", "사용자 김철수 계정 비활성화", "사원퇴사"),
    (2, "2026-05-21 10:22:45", "30일 이상 로그인 없음", "일정기간 로그인없음"),
    (3, "2026-05-22 08:30:00", "90일 이상 패스워드 변경 안함", "패스워드 변경안함"),
    (4, "2026-05-22 14:17:33", "사용자 박영희 계정 비활성화", "사원퇴사"),
    (5, "2026-05-23 07:00:10", "60일 이상 로그인 없음", "일정기간 로그인없음"),
], 1):
    ws_cumul.cell(row=idx+1, column=1, value=data[0])
    ws_cumul.cell(row=idx+1, column=2, value=data[1])
    ws_cumul.cell(row=idx+1, column=3, value=data[2])
    ws_cumul.cell(row=idx+1, column=4, value=data[3])

# AD외부망 시트
ws_ext = wb_cumul.create_sheet("AD외부망")
for col, h in enumerate(headers, 1):
    cell = ws_ext.cell(row=1, column=col, value=h)
    cell.font = bold
for idx, data in enumerate([
    (1, "2026-05-20 09:15:30", "사용자 김철수 계정 비활성화", "사원퇴사"),
    (2, "2026-05-22 14:17:33", "사용자 박영희 계정 비활성화", "사원 퇴사"),
], 1):
    ws_ext.cell(row=idx+1, column=1, value=data[0])
    ws_ext.cell(row=idx+1, column=2, value=data[1])
    ws_ext.cell(row=idx+1, column=3, value=data[2])
    ws_ext.cell(row=idx+1, column=4, value=data[3])

wb_cumul.save("/Users/scott/hermes room/ad_login_log/test_cumulative.xlsx")

# 원본 파일 (새로운 데이터, 중복 포함)
wb_orig = openpyxl.Workbook()
ws_orig = wb_orig.active
ws_orig.title = "원본"

# 헤더: no | 로그일자 | 내용 | 분류
orig_headers = ["no", "로그일자", "내용", "분류"]
for col, h in enumerate(orig_headers, 1):
    cell = ws_orig.cell(row=1, column=col, value=h)
    cell.font = bold

# 데이터 (누적 파일에 없는 새 데이터 + 중복)
for idx, data in enumerate([
    (None, "2026-05-24 11:30:00", "사용자 이영수 계정 비활성화", "사원퇴사"),
    (None, "2026-05-24 14:00:00", "90일 이상 패스워드 변경 안함", "패스워드 변경않함"),
    (None, "2026-05-24 15:45:00", "45일 이상 로그인 없음", "일정기간 로그인없음"),
    (None, "2026-05-25 08:00:00", "사용자 최자혜 계정 비활성화", "사원퇴사"),
    (None, "2026-05-25 09:30:00", "기타 사유", "기타"),
    # 중복 (누적 파일의 No 2와 동일한 내용)
    (None, "2026-05-21 10:22:45", "30일 이상 로그인 없음", "일정기간 로그인없음"),
    # 사원퇴사 중복 (누적에 있음)
    (None, "2026-05-20 09:15:30", "사용자 김철수 계정 비활성화", "사원퇴사"),
], 1):
    ws_orig.cell(row=idx+1, column=1, value=data[0])
    ws_orig.cell(row=idx+1, column=2, value=data[1])
    ws_orig.cell(row=idx+1, column=3, value=data[2])
    ws_orig.cell(row=idx+1, column=4, value=data[3])

wb_orig.save("/Users/scott/hermes room/ad_login_log/test_input.xlsx")

print("테스트 데이터 생성 완료!")
print("  - test_cumulative.xlsx (기존 No 1~5)")
print("  - test_input.xlsx (새 데이터 7개, 중복 2개, 스킵 1개)")
