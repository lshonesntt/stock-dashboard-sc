#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AD 로그인 로그 자동화 테스트 데이터 생성 스크립트"""

import openpyxl
from openpyxl.styles import Font, Alignment

def create_sample(filename, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AD로그"
    
    headers = ["no", "로그일자", "내용", "분류"]
    bold_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, values in enumerate(rows, 2):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
    
    col_widths = [10, 25, 40, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    wb.save(f"/Users/scott/hermes room/ad_login_log/{filename}")
    print(f"✓ {filename} 생성 완료")

# 샘플 1: 기본 패턴 (모든 사유 포함)
sample1_rows = [
    [None, "2026-05-24 09:00:00", "사용자 김철수 계정 비활성화", "사원퇴사"],
    [None, "2026-05-24 10:30:00", "90일 이상 로그인 없음", "일정기간 로그인없음"],
    [None, "2026-05-24 14:15:00", "180일 이상 패스워드 변경 안함", "패스워드 변경안함"],
    [None, "2026-05-25 08:00:00", "사용자 박영희 계정 비활성화", "사원퇴사"],
    [None, "2026-05-25 11:20:00", "60일 이상 로그인 없음", "일정기간 로그인없음"],
]
create_sample("sample1.xlsx", sample1_rows)

# 샘플 2: 중복 데이터 포함
sample2_rows = [
    [None, "2026-05-23 09:00:00", "사용자 이영수 계정 비활성화", "사원퇴사"],
    [None, "2026-05-23 10:00:00", "90일 이상 로그인 없음", "일정기간 로그인없음"],
    # 중복 데이터 (sample1과 동일한 내용)
    [None, "2026-05-24 09:00:00", "사용자 김철수 계정 비활성화", "사원퇴사"],
    [None, "2026-05-24 09:00:00", "사용자 김철수 계정 비활성화", "사원퇴사"],
    [None, "2026-05-25 15:00:00", "사용자 최자혜 계정 비활성화", "사원퇴사"],
    [None, "2026-05-25 16:30:00", "120일 이상 패스워드 변경 안함", "패스워드 변경안함"],
]
create_sample("sample2.xlsx", sample2_rows)

# 샘플 3: 예외 케이스 (허용되지 않은 사유)
sample3_rows = [
    [None, "2026-05-22 08:30:00", "사용자 정민수 계정 비활성화", "사원퇴사"],
    [None, "2026-05-22 09:15:00", "계정 잠김", "계정 잠김"],
    [None, "2026-05-22 10:00:00", "비밀번호 만료", "비밀번호 만료"],
    [None, "2026-05-22 11:30:00", "90일 이상 로그인 없음", "일정기간 로그인없음"],
    [None, "2026-05-22 13:45:00", "기타 사유", "기타"],
    [None, "2026-05-23 09:00:00", "120일 이상 패스워드 변경 안함", "패스워드 변경안함"],
    [None, "2026-05-23 10:30:00", "사용자 손예진 계정 비활성화", "사원퇴사"],
    [None, "2026-05-23 14:00:00", "사용자 황정민 계정 비활성화", "사원퇴사"],
]
create_sample("sample3.xlsx", sample3_rows)

print("\n완료: 3개 테스트 파일 생성")
print("   - sample1.xlsx: 기본 패턴 (5행)")
print("   - sample2.xlsx: 중복 데이터 포함 (6행, 중복 2개)")
print("   - sample3.xlsx: 예외 케이스 (8행, 스킵 3개)")
