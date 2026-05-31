#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IP 분류 도구 테스트 스크립트 - PyQt5 없이 독립 실행
.xlsx와 .xls 모두 테스트
"""

import sys
import os
from datetime import datetime

import openpyxl
import xlrd


def read_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row_num in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=row_num, column=col_idx + 1).value for col_idx in range(len(headers))]
        processed = []
        for val in row_vals:
            if isinstance(val, datetime):
                processed.append(val)
            elif isinstance(val, float):
                if val == int(val):
                    processed.append(int(val))
                else:
                    processed.append(val)
            else:
                processed.append(val)
        rows.append(processed)
    return headers, rows


def read_xls(filepath):
    book = xlrd.open_workbook(filepath)
    ws = book.sheet_by_index(0)
    headers = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]
    rows = []
    for row_idx in range(1, ws.nrows):
        row_vals = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
        processed = []
        for col_idx, val in enumerate(row_vals):
            ctype = ws.cell_type(row_idx, col_idx)
            if ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_tuple(val, book.datemode)
                    processed.append(datetime(*dt))
                except Exception:
                    processed.append(str(val))
            elif isinstance(val, float):
                if val == int(val):
                    processed.append(int(val))
                else:
                    processed.append(val)
            else:
                processed.append(val)
        rows.append(processed)
    return headers, rows


def format_timestamp(val):
    if not val:
        return ""
    if isinstance(val, str):
        return val
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def process_file(filepath, event_type="", enforcer="", ip_param="", admin="", timestamp=""):
    filepath_lower = filepath.lower()
    ext = os.path.splitext(filepath_lower)[1]

    if ext == ".xls":
        headers, rows_data = read_xls(filepath)
        print(f"   .xls 파일 읽기 완료: {len(rows_data)}행")
    else:
        headers, rows_data = read_xlsx(filepath)
        print(f"   .xlsx 파일 읽기 완료: {len(rows_data)}행")

    event_col = None
    for i, h in enumerate(headers):
        if h and "이벤트" in str(h):
            event_col = i
            print(f"   '이벤트 종류' 열 발견: col {event_col} = '{h}'")
            break

    if event_col is None:
        print("   '이벤트 종류' 열을 찾을 수 없습니다!")
        return None

    new_ip_rows = []
    del_ip_rows = []

    for row_idx, row_vals in enumerate(rows_data):
        if event_col >= len(row_vals):
            continue
        event_val = row_vals[event_col]
        if not event_val or str(event_val).strip() == "":
            continue
        event_str = str(event_val).strip()

        record = {}
        for h_idx, h in enumerate(headers):
            if h_idx < len(row_vals):
                record[str(h)] = row_vals[h_idx]
            else:
                record[str(h)] = ""

        if event_type and str(event_str) != event_type:
            continue
        if enforcer and str(record.get("엔포서명", "") or "") != enforcer:
            continue
        if ip_param and str(record.get("IP", "") or "") != ip_param:
            continue
        if admin and str(record.get("관리자", "") or "") != admin:
            continue
        if timestamp and str(record.get("발생일시", "") or "") != timestamp:
            continue

        if "IP 등록" in event_str or "인증허가" in event_str:
            new_ip_rows.append(record)
        elif "IP 삭제" in event_str or "일정기간 미사용 IP삭제" in event_str:
            del_ip_rows.append(record)

    print(f"  분류 결과:")
    print(f"     - 신규IP: {len(new_ip_rows)}행")
    print(f"     - IP삭제: {len(del_ip_rows)}행")

    if new_ip_rows:
        print(f"\n   신규IP 행 목록:")
        for idx, rec in enumerate(new_ip_rows, 1):
            ts = format_timestamp(rec.get("발생일시", ""))
            print(f"     {idx}. [{ts}] {rec.get('엔포서명', '')} | {rec.get('IP', '')} | {rec.get('관리자', '')} | {rec.get('이벤트 종류', '')}")

    if del_ip_rows:
        print(f"\n   IP삭제 행 목록:")
        for idx, rec in enumerate(del_ip_rows, 1):
            ts = format_timestamp(rec.get("발생일시", ""))
            print(f"     {idx}. [{ts}] {rec.get('엔포서명', '')} | {rec.get('IP', '')} | {rec.get('관리자', '')} | {rec.get('이벤트 종류', '')}")

    return {
        "new_ip": new_ip_rows,
        "del_ip": del_ip_rows,
    }


def save_output(original_path, new_ip_rows, del_ip_rows, output_dir="."):
    out_wb = openpyxl.Workbook()
    headers = ["No", "발생일시", "엔포서명", "IP", "관리자", "이벤트 종류"]
    bold_font = openpyxl.styles.Font(bold=True)
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin"),
        right=openpyxl.styles.Side(style="thin"),
        top=openpyxl.styles.Side(style="thin"),
        bottom=openpyxl.styles.Side(style="thin"),
    )

    for ws_title, rows in [("신규IP", new_ip_rows), ("IP삭제", del_ip_rows)]:
        ws = out_wb.active if ws_title == "신규IP" else out_wb.create_sheet(ws_title)
        if ws_title == "IP삭제":
            out_wb.move_sheet(ws_title, offset=1)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        for idx, row_data in enumerate(rows, 1):
            ws.cell(row=idx + 1, column=1, value=idx).border = thin_border
            ws.cell(row=idx + 1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")
            ws.cell(row=idx + 1, column=2, value=format_timestamp(row_data.get("발생일시", ""))).border = thin_border
            ws.cell(row=idx + 1, column=3, value=row_data.get("엔포서명", "") or "").border = thin_border
            ws.cell(row=idx + 1, column=4, value=str(row_data.get("IP", "") or "")).border = thin_border
            ws.cell(row=idx + 1, column=5, value=row_data.get("관리자", "") or "").border = thin_border
            ws.cell(row=idx + 1, column=6, value=str(row_data.get("이벤트 종류", "") or "")).border = thin_border

    max_widths = [10] * len(headers)
    for rows in [new_ip_rows, del_ip_rows]:
        for idx, row_data in enumerate(rows, 1):
            vals = [
                str(idx),
                format_timestamp(row_data.get("발생일시", "")),
                str(row_data.get("엔포서명", "") or ""),
                str(row_data.get("IP", "") or ""),
                str(row_data.get("관리자", "") or ""),
                str(row_data.get("이벤트 종류", "") or ""),
            ]
            for col_idx, val in enumerate(vals):
                w = len(val.encode("utf-8"))
                if w > max_widths[col_idx]:
                    max_widths[col_idx] = min(w, 60)

    for ws in [out_wb.worksheets[0], out_wb.worksheets[1]]:
        for col_idx, max_w in enumerate(max_widths):
            col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
            ws.column_dimensions[col_letter].width = max_w + 4

    out_name = os.path.splitext(os.path.basename(original_path))[0] + "_classified.xlsx"
    output_path = os.path.join(output_dir if output_dir else ".", out_name)
    out_wb.save(output_path)
    print(f"\n   출력 파일 저장됨: {output_path}")
    return output_path


# ── 테스트 실행 ──────────────────────────────────────────────────────

SAMPLE_DIR = "/Users/scott/hermes room/ip_classifier"

print("=" * 70)
print(" IP 분류 도구 테스트")
print("=" * 70)

# 1. xlsx 파일 테스트
print(f"\n{'='*70}")
print(" 1. .xlsx 파일 테스트")
print(f"{'='*70}")

xlsx_files = [
    f"{SAMPLE_DIR}/samples_1.xlsx",
    f"{SAMPLE_DIR}/samples_2.xlsx",
    f"{SAMPLE_DIR}/samples_3.xlsx",
]

for fp in xlsx_files:
    print(f"\n📄 {os.path.basename(fp)}:")
    results = process_file(fp)
    if results and results["new_ip"] and results["del_ip"]:
        save_output(fp, results["new_ip"], results["del_ip"], SAMPLE_DIR)

# 2. xls 파일 테스트
print(f"\n{'='*70}")
print(" 2. .xls 파일 테스트")
print(f"{'='*70}")

xls_files = [
    f"{SAMPLE_DIR}/samples_1.xls",
]

for fp in xls_files:
    print(f"\n📄 {os.path.basename(fp)}:")
    results = process_file(fp)
    if results and results["new_ip"] and results["del_ip"]:
        save_output(fp, results["new_ip"], results["del_ip"], SAMPLE_DIR)

# 3. 파라미터 필터링 테스트
print(f"\n{'='*70}")
print(" 3. 파라미터 필터링 테스트 (samples_1.xlsx)")
print(f"{'='*70}")
results = process_file(f"{SAMPLE_DIR}/samples_1.xlsx", enforcer="firewall01")

print(f"\n{'='*70}")
print("✅ 전체 테스트 완료!")
print(f"{'='*70}")
